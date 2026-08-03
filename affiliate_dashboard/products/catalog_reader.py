"""Liest einen Produkt-Katalog-Tab (z. B. "tauchpumpe05-21", "tischkreissaege0a-21")
aus dem Google Sheet -- eine Zeile pro getestetem Produkt.

Andere Semantik als der Umsatz-Import (``adapters/gsheet_adapter.py``): hier ist
jede Zeile ein Produkt mit Name/ASIN/Test-URL/Amazon-URL/Kategorie plus frei
variierenden technischen Spezifikationen (je Nische unterschiedliche Spalten)
-- die Spezifikationen werden generisch als ``{header: wert}``-Dict
mitgenommen, nicht in ein festes Schema gezwungen. Die erste Spalte jedes Tabs
traegt als Kopfzeile den Namen der Nische selbst und wird deshalb positionell
(Index 0) als Produktname behandelt, nicht ueber Alias-Matching.

Die Tracking-ID wird NICHT aus einer Spalte gelesen (nicht jeder Tab hat eine
"Tracking ID"-Spalte -- z. B. der Tischkreissaegen-Tab), sondern kommt 1:1 vom
konfigurierten Tab-Label (``/settings`` -> Produkt-Tabs). Wer diese Zuordnung
nutzt, muss das Tab-Label exakt auf die echte Tracking-ID setzen (z. B.
"tauchpumpe05-21").

Gelesen wird als **XLSX**, nicht CSV: manche Tabs (z. B. Tischkreissaegen)
haben keine eigene "Test"-Spalte, sondern verlinken den Testartikel direkt auf
der Produktname-Zelle als Hyperlink -- ein CSV-Export wuerde diesen Link
verlieren (CSV kennt nur sichtbaren Text), der XLSX-Export traegt ihn. Die
"Test"-Spalte hat weiterhin Vorrang, falls vorhanden und befuellt; der
Hyperlink auf der Namenszelle ist nur der Fallback.
"""

from __future__ import annotations

import io

import openpyxl

from .. import columns
from ..gsheet_fetch import fetch_xlsx


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get(values: list[str], idx: int | None) -> str:
    if idx is None or idx >= len(values):
        return ""
    return values[idx]


def read_tab(sheet_id: str, gid: str, tab_label: str) -> tuple[list[dict], dict]:
    """Einen Produkt-Tab lesen. Gibt ``(items, stats)`` zurueck.

    ``items``: Liste von ``{tab_label, name, tracking_id, asin, category,
    test_url, amazon_url, specs}`` -- ``tracking_id`` ist immer gleich
    ``tab_label``. Zeilen ohne ASIN werden uebersprungen -- ohne ASIN ist
    weder eine Amazon-Verfuegbarkeitspruefung noch eine
    Besucherzahlen-Zuordnung moeglich.
    """
    data = fetch_xlsx(sheet_id, gid)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(min_row=1)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        return [], {"rows": 0, "no_asin": 0}
    headers = [_cell_str(c.value) for c in header_cells]

    cols = columns.resolve_columns(headers, aliases=columns.PRODUCT_ALIASES)
    category_idx = cols.get("category")
    specs_start = (category_idx + 1) if category_idx is not None else len(headers)

    items: list[dict] = []
    stats = {"rows": 0, "no_asin": 0}

    for row_cells in rows_iter:
        values = [_cell_str(c.value) for c in row_cells]
        if not any(values):
            continue  # komplett leere Zeile

        name = _get(values, 0)
        asin = _get(values, cols.get("asin"))
        if not name and not asin:
            continue  # keine sinnvolle Produktzeile

        stats["rows"] += 1
        if not asin:
            stats["no_asin"] += 1
            continue

        test_url = _get(values, cols.get("test_url"))
        if not test_url and row_cells and row_cells[0].hyperlink:
            test_url = row_cells[0].hyperlink.target or ""

        specs: dict[str, str] = {}
        for idx in range(specs_start, len(headers)):
            header = headers[idx]
            if not header:
                continue
            value = _get(values, idx)
            if value:
                specs[header] = value

        items.append({
            "tab_label": tab_label,
            "name": name,
            "tracking_id": tab_label,
            "asin": asin,
            "category": _get(values, cols.get("category")),
            "test_url": test_url,
            "amazon_url": _get(values, cols.get("amazon_url")),
            "specs": specs,
        })

    return items, stats
