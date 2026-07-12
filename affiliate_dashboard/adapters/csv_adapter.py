"""CSV-/XLSX-Datei-Adapter (generischer Zusatzweg).

Liest alle Dateien aus ``data/inbox`` (``*.csv``, ``*.txt``, ``*.xlsx``),
normalisiert sie und verschiebt sie nach ``data/archive/<datum>/``.
Gedacht fuer einmalige PartnerNet-Exporte abseits des Google Sheets.

PartnerNet-CSV ist oft TAB-getrennt und als ``.txt`` exportiert -> der
Delimiter wird per ``csv.Sniffer`` erkannt (Fallback: Tab, dann Komma).
``.xlsx`` benoetigt ``openpyxl`` (optional, lazy importiert).
"""

from __future__ import annotations

import csv
import shutil
from datetime import datetime
from pathlib import Path

from .base import IngestionAdapter, IngestResult
from .. import columns

_TEXT_SUFFIXES = {".csv", ".txt", ".tsv"}


class CsvAdapter(IngestionAdapter):
    source = "csv"

    def load(self) -> IngestResult:
        inbox = self.config.path("inbox_dir")
        inbox.mkdir(parents=True, exist_ok=True)
        files = sorted(
            p for p in inbox.iterdir()
            if p.is_file() and p.suffix.lower() in (_TEXT_SUFFIXES | {".xlsx"})
        )
        if not files:
            raise ValueError(
                f"Keine Dateien in {inbox} gefunden. Lege einen PartnerNet-Export "
                f"(.csv/.txt/.xlsx) dort ab."
            )

        all_items: list[dict] = []
        agg_stats = {"rows": 0, "no_tracking_id": 0, "no_month": 0, "files": 0}
        processed: list[Path] = []

        for path in files:
            headers, rows = self._read_file(path)
            # Monat ggf. aus Dateiname ableiten (z. B. "...-2024-03.csv")
            default_month = columns.parse_month(_month_from_name(path.stem))
            items, stats = columns.normalize_rows(headers, rows, default_month=default_month)
            all_items.extend(items)
            agg_stats["files"] += 1
            for k in ("rows", "no_tracking_id", "no_month"):
                agg_stats[k] += stats.get(k, 0)
            processed.append(path)

        self._archive(processed)
        return IngestResult(
            items=all_items,
            source=self.source,
            source_file=", ".join(p.name for p in processed),
            full_snapshot=False,  # einzelne Dateien -> inkrementell upserten
            stats=agg_stats,
        )

    # --- intern -----------------------------------------------------------
    def _read_file(self, path: Path):
        if path.suffix.lower() == ".xlsx":
            return self._read_xlsx(path)
        return self._read_text(path)

    def _read_text(self, path: Path):
        raw = path.read_bytes()
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if "\t" in sample else ","

        import io
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]

    def _read_xlsx(self, path: Path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Fuer .xlsx-Import wird 'openpyxl' benoetigt: pip install openpyxl"
            ) from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[("" if c is None else c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        if not rows:
            return [], []
        return rows[0], rows[1:]

    def _archive(self, files: list[Path]) -> None:
        archive = self.config.path("archive_dir") / datetime.now().strftime("%Y-%m-%d")
        archive.mkdir(parents=True, exist_ok=True)
        for path in files:
            try:
                shutil.move(str(path), str(archive / path.name))
            except Exception:
                pass


def _month_from_name(stem: str) -> str:
    import re
    m = re.search(r"(\d{4}[-_.]\d{1,2})", stem)
    return m.group(1).replace("_", "-").replace(".", "-") if m else ""
